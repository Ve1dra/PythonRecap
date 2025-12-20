import openpyxl
from openpyxl import load_workbook
# from time import ctime
def calender():pass

def schedule():
    file = "scheduleData.xlsx"
    wb = load_workbook(filename=file)
    sheet = wb.active
    # sheet['D1'] = 'Wednesday'
    # sheet.cell(row=2, column=2, value='Another Value')
    a=1
    while a:
        print("Enter 'done' to stop entering events\nEnter 'update' to update your schedule\nEnter 'delete' to deleat an entry.")

        ent = str(input("Enter the operation you plan to do: "))
        if ent == "done".lower():
            break

        elif ent == "update".lower():
            u_row = int(input("Enter the row you wish to update: "))
            # sheet.delete_rows(u_row)
            # new_row = sheet.insert_rows(u_row, 1)
            u_event = input("Enter the new event: ")
            u_time =  str(input("Enter the time this event would hold: "))
            print("\n")
            # u_row_data = [u_time, u_event]
            sheet.cell(u_row, 1, u_time)
            sheet.cell(u_row, 2, u_event)
            wb.save('scheduleData.xlsx')

        elif ent == "delete".lower():
            d_row = int(input("Enter the row you want to delete: "))
            sheet.delete_rows(d_row)
            wb.save('scheduleData.xlsx')

        elif '0' <= ent <= '9':
            print("Invalid Input")

        else:
            event = ent
            day = str(input("Enter the time this event would hold: "))
            print("\n")
            time = day
            new_row_data = [time, event]
            sheet.append(new_row_data)
            wb.save("scheduleData.xlsx")

    for row in sheet.iter_rows(values_only=True):
        print("\t".join([str(cell) if cell is not None  else "" for cell in row]))
    wb.close()
    